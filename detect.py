import time
import random    
import tkinter as tk
import cv2
import math
import argparse
import os
import PIL

import tkinter as tk
from tkinter import messagebox 
from tkinter import *
from openpyxl import *
from PIL import ImageTk,Image

def show_image():
    if check_btn_var1.get() == 1:
        image_1label.config(image=tsimage)
    else:
        image_1label.config(image="")
    
    if check_btn_var2.get() == 1:
        image_2label.config(image=msimage)
    else:
        image_2label.config(image="")

    if check_btn_var3.get() == 1:
        image_3label.config(image=sleevlessimage)
    else:
        image_3label.config(image="")

    if check_btn_var4.get() == 1:
        image_4label.config(image=suitimage)
    else:
        image_4label.config(image="")

    if check_btn_var5.get() == 1:
        image_5label.config(image=sweaterimg)
    else:
        image_5label.config(image="")

    if check_btn_var6.get() == 1:
        image_6label.config(image=cardiganimg)
    else:
        image_6label.config(image="")
    
    if lf2_check_btn_var7.get() == 1:
        image_7label.config(image=piqueimg)
    else:
        image_7label.config(image="")

    if check_btn_var8.get() == 1:
        image_8label.config(image=vimage)
    else:
        image_8label.config(image="")
    
    if check_btn_var9.get() == 1:
        image_9label.config(image=capeimg)
    else:
        image_9label.config(image="")

    if check_btn_var10.get() == 1:
        image_10label.config(image=coatimg)
    else:
        image_10label.config(image="")
    
    if check_btn_var11.get() == 1:
        image_11label.config(image=blouseimg)
    else:
        image_11label.config(image="")

    if check_btn_var12.get() == 1:
        image_12label.config(image=sweatsuitimg)
    else:
        image_12label.config(image="")


##남자 하의
    if  lf2_check_btn_var1.get() == 1:
        bimage_1label.config(image=Jeansimg)
    else:
        bimage_1label.config(image="")
    
    if lf2_check_btn_var2.get() == 1:
        bimage_2label.config(image=Shortsimg)
    else:
        bimage_2label.config(image="")

    if lf2_check_btn_var3.get() == 1:
        bimage_3label.config(image=Straight_wool_trouserimg)
    else:
        bimage_3label.config(image="")

    if lf2_check_btn_var4.get() == 1:
        bimage_4label.config(image=Dungareeimg)
    else:
        bimage_4label.config(image="")

    if lf2_check_btn_var5.get() == 1:
        bimage_5label.config(image=Taylored_shortsimg)
    else:
        bimage_5label.config(image="")

    if lf2_check_btn_var6.get() == 1:
        bimage_6label.config(image=Elastic_trouserimg)
    else:
        bimage_6label.config(image="")
    
    if lf2_check_btn_var7.get() == 1:
        bimage_7label.config(image=Towering_shortsimg)
    else:
        bimage_7label.config(image="")

    if lf2_check_btn_var8.get() == 1:
        bimage_8label.config(image=Relaxedfit_jeansimg)
    else:
        bimage_8label.config(image="")
    
    if lf2_check_btn_var9.get() == 1:
        bimage_9label.config(image=uility_shortsimg)
    else:
        bimage_9label.config(image="")

    if lf2_check_btn_var10.get() == 1:
        bimage_10label.config(image=Seersucker_shortsimg)
    else:
        bimage_10label.config(image="")
    
    if lf2_check_btn_var11.get() == 1:
        bimage_11label.config(image=cargo_pantsimg)
    else:
        bimage_11label.config(image="")

    if lf2_check_btn_var12.get() == 1:
        bimage_12label.config(image=swarm_shortsimg)
    else:
        bimage_12label.config(image="")


def error_handling(msg):
    if msg == "error!":
        messagebox.showerror("Error")
    elif msg == "division by zero!":
        messagebox.showerror("Error")

def exit_program():
    root.destroy()



def reset_sales():
    try:
        check_btn_var1.set(0)
        check_btn_var2.set(0)
        check_btn_var3.set(0)
        check_btn_var4.set(0)
        check_btn_var5.set(0)
        check_btn_var6.set(0)
        check_btn_var7.set(0)
        check_btn_var8.set(0)
        check_btn_var9.set(0)
        check_btn_var10.set(0)
        check_btn_var11.set(0)
        check_btn_var12.set(0)

        global list_entry
        entry_opt_var1.set(list_entry[0])
        entry_opt_var2.set(list_entry[0])
        entry_opt_var3.set(list_entry[0])
        entry_opt_var4.set(list_entry[0])
        entry_opt_var5.set(list_entry[0])
        entry_opt_var6.set(list_entry[0])
        entry_opt_var7.set(list_entry[0])
        entry_opt_var8.set(list_entry[0])
        entry_opt_var9.set(list_entry[0])
        entry_opt_var10.set(list_entry[0])
        entry_opt_var11.set(list_entry[0])
        entry_opt_var12.set(list_entry[0])
     
        lf2_check_btn_var1.set(0)
        lf2_check_btn_var2.set(0)
        lf2_check_btn_var3.set(0)
        lf2_check_btn_var4.set(0)
        lf2_check_btn_var5.set(0)
        lf2_check_btn_var6.set(0)
        lf2_check_btn_var7.set(0)
        lf2_check_btn_var8.set(0)
        lf2_check_btn_var9.set(0)
        lf2_check_btn_var10.set(0)
        lf2_check_btn_var11.set(0)
        lf2_check_btn_var12.set(0)

        image_1label.config(image="")
        image_2label.config(image="")
        image_3label.config(image="")
        image_4label.config(image="")
        image_5label.config(image="")
        image_6label.config(image="")
        image_7label.config(image="")
        image_8label.config(image="")
        image_9label.config(image="")
        image_10label.config(image="")
        image_11label.config(image="")
        image_12label.config(image="")

        bimage_1label.config(image="")
        bimage_2label.config(image="")
        bimage_3label.config(image="")
        bimage_4label.config(image="")
        bimage_5label.config(image="")
        bimage_6label.config(image="")
        bimage_7label.config(image="")
        bimage_8label.config(image="")
        bimage_9label.config(image="")
        bimage_10label.config(image="")
        bimage_11label.config(image="")
        bimage_12label.config(image="")

        global lf2_list_entry
        lf2_entry_opt_var1.set(lf2_list_entry[0])
        lf2_entry_opt_var2.set(lf2_list_entry[0])
        lf2_entry_opt_var3.set(lf2_list_entry[0])
        lf2_entry_opt_var4.set(lf2_list_entry[0])
        lf2_entry_opt_var5.set(lf2_list_entry[0])
        lf2_entry_opt_var6.set(lf2_list_entry[0])
        lf2_entry_opt_var7.set(lf2_list_entry[0])
        lf2_entry_opt_var8.set(lf2_list_entry[0])
        lf2_entry_opt_var9.set(lf2_list_entry[0])
        lf2_entry_opt_var10.set(lf2_list_entry[0])
        lf2_entry_opt_var11.set(lf2_list_entry[0])
        lf2_entry_opt_var12.set(lf2_list_entry[0])

        entry_cost_top.delete(0, END)
        entry_cost_bottom.delete(0, END)
        entry_total_cost.delete(0, END)
        tax_entry.delete(0, END)
        subtotal_entry.delete(0, END)
        total_entry.delete(0, END)
        clothes_field.delete("1.0", END)
        clothes_quantity_price.clear()
       
        
        messagebox.showinfo("thanks")
    except NameError:
        messagebox.showerror("Error!", "Nothing to clear!")



def total_salesM():
    # ================================ top ============================== #
    price_tagM = {"T_shirt": 20000, "Shirt": 10000, "Sleevless_shirt": 25000, "Suit": 100000, "Sweater": 120000, "Cardigan": 110000, "Pique_shirt": 30000,
                 "Vest": 35000, "Cape": 20000, "Coat": 450000, "Blouse": 25000, "Sweat_shirt": 30000,
                }

    global clothes_quantity_price
    clothes_quantity_price = []

    if check_btn_var1.get() == 1:
        T_shirt_results = int(entry_opt_var1.get()) * price_tagM["T_shirt"]
        T_shirt_tuple = ("티셔츠", str(entry_opt_var1.get()), "20000")
        clothes_quantity_price.append(T_shirt_tuple)
        image_1label.config(image=tsimage)
        image_1label.place()
    else:
        T_shirt_results = 0

    if check_btn_var2.get() == 1:
        Shirt_results = int(entry_opt_var2.get()) * price_tagM[ "Shirt"]
        Shirt_tuple = ("셔츠", str(entry_opt_var2.get()), "10000")
        clothes_quantity_price.append(Shirt_tuple)
        image_2label.config(image=msimage)
        image_2label.place()
    else:
        Shirt_results = 0

    if check_btn_var3.get() == 1:
        Sleeveless_shirt_results = int(entry_opt_var3.get()) * price_tagM["Sleevless_shirt"]
        Sleevless_shirt_tuple = ("민소매", str(entry_opt_var3.get()), "25000")
        clothes_quantity_price.append(Sleevless_shirt_tuple)
        image_3label.config(image=sleevlessimage)
        image_3label.place()
    else:
        Sleeveless_shirt_results = 0

    if check_btn_var4.get() == 1:
        Suit_results = int(entry_opt_var4.get()) * price_tagM["Suit"]
        Suit_tuple = ("정장", str(entry_opt_var4.get()), "100000")
        clothes_quantity_price.append(Suit_tuple)
        image_4label.config(image=suitimage)
        image_4label.place()
    else:
        Suit_results = 0

    if check_btn_var5.get() == 1:
        Sweater_results = int(entry_opt_var5.get()) * price_tagM["Sweater"]
        Sweater_tuple = ("스웨터", str(entry_opt_var5.get()), "120000")
        clothes_quantity_price.append(Sweater_tuple)
        image_5label.config(image=sweaterimg)
        image_5label.place()
    else:
        Sweater_results = 0

    if check_btn_var6.get() == 1:
        Cardigan_results = int(entry_opt_var6.get()) * price_tagM["Cardigan"]
        Cardigan_tuple = ("가디건", str(entry_opt_var6.get()), "110000")
        clothes_quantity_price.append(Cardigan_tuple)
        image_6label.config(image=cardiganimg)
        image_6label.place()
    else:
        Cardigan_results = 0

    if check_btn_var7.get() == 1:
        Pique_shirt_results = int(entry_opt_var7.get()) * price_tagM["Pique_shirt"]
        Pique_shirt_tuple = ("피케셔츠", str(entry_opt_var7.get()), "30000")
        clothes_quantity_price.append(Pique_shirt_tuple)
        image_7label.config(image=piqueimg)
        image_7label.place()
    else:
        Pique_shirt_results = 0

    if check_btn_var8.get() == 1:
        Vest_results = int(entry_opt_var8.get()) * price_tagM["Vest"]
        Vest_tuple = ("조끼", str(entry_opt_var8.get()), "35000")
        clothes_quantity_price.append(Vest_tuple)
        image_8label.config(image=vimage)
        image_8label.place()
    else:
        Vest_results = 0

    if check_btn_var9.get() == 1:
        Cape_results = int(entry_opt_var9.get()) * price_tagM["Cape"]
        Sleeveless_top_tuple = ("케이프 코트", str(entry_opt_var9.get()), "20000")
        clothes_quantity_price.append(Sleeveless_top_tuple)
        image_9label.config(image=capeimg)
        image_9label.place()
    else:
        Cape_results = 0

    if check_btn_var10.get() == 1:
        Coat_results = int(entry_opt_var10.get()) * price_tagM["Coat"]
        Coat_tuple = ("코트", str(entry_opt_var10.get()), "450000")
        clothes_quantity_price.append(Coat_tuple)
        image_10label.config(image=coatimg)
        image_10label.place()
    else:
        Coat_results = 0

    if check_btn_var11.get() == 1:
        Blouse_results = int(entry_opt_var11.get()) * price_tagM["Blouse"]
        Blouse_tuple = ("블라우스", str(entry_opt_var11.get()), "25000")
        clothes_quantity_price.append(Blouse_tuple)
        image_11label.config(image=blouseimg)
        image_11label.place()
    else:
        Blouse_results = 0

    if check_btn_var12.get() == 1:
        Sweat_shirt_results = int(entry_opt_var12.get()) * price_tagM["Sweat_shirt"]
        Sweat_shirt_tuple = ("운동복", str(entry_opt_var12.get()), "Sweat_shirt")
        clothes_quantity_price.append(Sweat_shirt_tuple)
        image_12label.config(image=sweatsuitimg)
        image_12label.place()
    else:
        Sweat_shirt_results = 0

    try:
        global total_cost_of_top
        

        total_cost_of_top = (T_shirt_results + Shirt_results + Sleeveless_shirt_results + Suit_results + Sweater_results +
                                Cardigan_results + Pique_shirt_results + Vest_results + Cape_results +
                                Coat_results + Blouse_results + Sweat_shirt_results    )

        entry_cost_top.delete(0, END)
        entry_cost_top.insert(0,  str(total_cost_of_top)+" 원. ")
    except NameError:
        error_handling("error")
    except ValueError:
        messagebox.showerror("Error")

    # ================================ bottom-top ==============================

    bottom_price_tagM = {"Jeans": 25000, "Shorts": 10500, "Straight_wool_trouser": 100000, "Dungaree": 175000,
                       "Taylored_shorts": 100000, "Elastic_trouser":115000, "Towering_shorts": 79000, "Relaxedfit_jeans": 12000,
                       "uility_shorts": 25000, "Seersucker_shorts": 100000, "cargo_pants": 180000, "swarm_shorts": 200000}

    if lf2_check_btn_var1.get() == 1:
        x1 = int(lf2_entry_opt_var1.get()) * bottom_price_tagM["Jeans"]
        Jeans_tuple = ("청바지", str(lf2_entry_opt_var1.get()), "25000")
        clothes_quantity_price.append(Jeans_tuple)
        bimage_1label.config(image=Jeansimg)
        bimage_1label.place()
    else:
        x1 = 0

    if lf2_check_btn_var2.get() == 1:
        x2 = int(lf2_entry_opt_var2.get()) * bottom_price_tagM["Shorts"]
        Shorts_tuple = ("반바지", str(lf2_entry_opt_var2.get()), "10500")
        clothes_quantity_price.append(Shorts_tuple)
        bimage_2label.config(image=Shortsimg)
        bimage_2label.place()
    else:
        x2 = 0

    if lf2_check_btn_var3.get() == 1:
        x3 = int(lf2_entry_opt_var3.get()) * bottom_price_tagM["Straight_wool_trouser"]
        Straight_wool_trouser_tuple = ("울바지", str(lf2_entry_opt_var3.get()), "100000")
        clothes_quantity_price.append(Straight_wool_trouser_tuple)
        bimage_3label.config(image=Straight_wool_trouserimg)
        bimage_3label.place()
    else:
        x3 = 0

    if lf2_check_btn_var4.get() == 1:
        x4 = int(lf2_entry_opt_var4.get()) * bottom_price_tagM["Dungaree"]
        Dungaree_tuple = ("멜빵바지", str(lf2_entry_opt_var4.get()), "175000")
        clothes_quantity_price.append(Dungaree_tuple)
        bimage_4label.config(image=Dungareeimg)
        bimage_4label.place()
    else:
        x4 = 0

    if lf2_check_btn_var5.get() == 1:
        x5 = int(lf2_entry_opt_var5.get()) * bottom_price_tagM["Taylored_shorts"]
        Taylored_shorts_tuple = ("정장 바지", str(lf2_entry_opt_var5.get()), "100000")
        clothes_quantity_price.append(Taylored_shorts_tuple)
        bimage_5label.config(image=Taylored_shortsimg)
        bimage_5label.place()
    else:
        x5 = 0

    if lf2_check_btn_var6.get() == 1:
        x6 = int(lf2_entry_opt_var6.get()) * bottom_price_tagM["Elastic_trouser"]
        Elastic_trouser_tuple = ("기능성 바지", str(lf2_entry_opt_var6.get()), "115000")
        clothes_quantity_price.append(Elastic_trouser_tuple)
        bimage_6label.config(image=Elastic_trouserimg)
        bimage_6label.place()
    else:
        x6 = 0

    if lf2_check_btn_var7.get() == 1:
        x7 = int(lf2_entry_opt_var7.get()) * bottom_price_tagM["Towering_shorts"]
        Towering_shorts_tuple = ("타월 바지", str(lf2_entry_opt_var7.get()), "79000")
        clothes_quantity_price.append(Towering_shorts_tuple)
        bimage_7label.config(image=Towering_shortsimg)
        bimage_7label.place()
    else:
        x7 = 0

    if lf2_check_btn_var8.get() == 1:
        x8 = int(lf2_entry_opt_var8.get()) * bottom_price_tagM["Relaxedfit_jeans"]
        Relaxedfit_jeans_tuple = ("릴렉스 핏 청바지", str(lf2_entry_opt_var8.get()), "12000")
        clothes_quantity_price.append(Relaxedfit_jeans_tuple)
        bimage_8label.config(image=Relaxedfit_jeansimg)
        bimage_8label.place()
    else:
        x8 = 0

    if lf2_check_btn_var9.get() == 1:
        x9 = int(lf2_entry_opt_var9.get()) * bottom_price_tagM["uility_shorts"]
        uility_shorts_tuple = ("운동바지", str(lf2_entry_opt_var9.get()), "25000")
        clothes_quantity_price.append(uility_shorts_tuple)
        bimage_9label.config(image=uility_shortsimg)
        bimage_9label.place()
    else:
        x9 = 0

    if lf2_check_btn_var10.get() == 1:
        x10 = int(lf2_entry_opt_var10.get()) * bottom_price_tagM["Seersucker_shorts"]
        Seersucker_shorts_tuple = ("시어서커 반바지", str(lf2_entry_opt_var10.get()), "100000")
        clothes_quantity_price.append(Seersucker_shorts_tuple)
        bimage_10label.config(image=Seersucker_shortsimg)
        bimage_10label.place()
    else:
        x10 = 0

    if lf2_check_btn_var11.get() == 1:
        x11 = int(lf2_entry_opt_var11.get()) * bottom_price_tagM["cargo_pants"]
        cargo_pants_tuple = ("카고바지", str(lf2_entry_opt_var11.get()), "180000")
        clothes_quantity_price.append(cargo_pants_tuple)
        bimage_11label.config(image=cargo_pantsimg)
        bimage_11label.place()
    else:
        x11 = 0

    if lf2_check_btn_var12.get() == 1:
        x12 = int(lf2_entry_opt_var12.get()) * bottom_price_tagM["swarm_shorts"]
        swarm_shorts_tuple = ("수영복 바지", str(lf2_entry_opt_var12.get()), "200000")
        clothes_quantity_price.append(swarm_shorts_tuple)
        bimage_12label.config(image=swarm_shortsimg)
        bimage_12label.place()
    else:
        x12 = 0

    try:
        global x13
      
        x13 = 0

        global total_cost_of_bottom
        global tax_on_all_items
        total_cost_of_bottom = (x1 + x2 + x3 + x4 + x5 + x6 + x7 + x8 + x9 + x10 + x11 + x12 + x13)

        entry_cost_bottom.delete(0, END)
        entry_cost_bottom.insert(0,  str(total_cost_of_bottom)+" 원. ")

        entry_total_cost.delete(0, END)
        entry_total_cost.insert(0, str(total_cost_of_bottom + total_cost_of_top)+" 원. ")

        tax_on_all_items = 0.10 * (total_cost_of_top + total_cost_of_bottom)

        tax_entry.delete(0, END)
        tax_entry.insert(0,  str(tax_on_all_items)+ " 원. " )

        subtotal_entry.delete(0, END)
        subtotal_entry.insert(0, str(total_cost_of_bottom + total_cost_of_top)+ " 원. ")

        total_entry.delete(0, END)
        total_entry.insert(0, str(total_cost_of_bottom + total_cost_of_top + tax_on_all_items)+" 원. ")
    except NameError:
        error_handling("error!")
    except ValueError:
        messagebox.showerror("Error!", "Price your items first, before finding the total.")

def total_salesw():
    # ================================ top ============================== #
    price_tagW = {"T_shirt": 20000, "Shirt": 13000, "Sleevless_shirt": 25000, "dress": 120000, "Sweater": 120000, "Cardigan": 110000, "Pique_shirt": 30000,
                 "WESTcoat": 135000, "Tanktop": 49000, "Coat": 450000, "Pencidress": 290000, "Maxidress": 390000,
                }

    global clothes_quantity_price
    clothes_quantity_price = []

    if check_btn_var1.get() == 1:
        T_shirt_results = int(entry_opt_var1.get()) * price_tagW["T_shirt"]
        T_shirt_tuple = ("티셔츠", str(entry_opt_var1.get()), "20000")
        clothes_quantity_price.append(T_shirt_tuple)
       
    else:
        T_shirt_results = 0

    if check_btn_var2.get() == 1:
        Shirt_results = int(entry_opt_var2.get()) * price_tagW[ "Shirt"]
        Shirt_tuple = ("셔츠", str(entry_opt_var2.get()), "13000")
        clothes_quantity_price.append(Shirt_tuple)
    else:
        Shirt_results = 0

    if check_btn_var3.get() == 1:
        Sleeveless_shirt_results = int(entry_opt_var3.get()) * price_tagW["Sleevless_shirt"]
        Sleevless_shirt_tuple = ("민소매 셔츠", str(entry_opt_var3.get()), "25000")
        clothes_quantity_price.append(Sleevless_shirt_tuple)
    else:
        Sleeveless_shirt_results = 0

    if check_btn_var4.get() == 1:
        dress_results = int(entry_opt_var4.get()) * price_tagW["dress"]
        dress_tuple = ("드레스", str(entry_opt_var4.get()), "120000")
        clothes_quantity_price.append(dress_tuple)
    else:
        dress_results = 0

    if check_btn_var5.get() == 1:
        Sweater_results = int(entry_opt_var5.get()) * price_tagW["Sweater"]
        Sweater_tuple = ("스웨터", str(entry_opt_var5.get()), "120000")
        clothes_quantity_price.append(Sweater_tuple)
    else:
        Sweater_results = 0

    if check_btn_var6.get() == 1:
        Cardigan_results = int(entry_opt_var6.get()) * price_tagW["Cardigan"]
        Cardigan_tuple = ("가디건", str(entry_opt_var6.get()), "110000")
        clothes_quantity_price.append(Cardigan_tuple)
    else:
        Cardigan_results = 0

    if check_btn_var7.get() == 1:
        Pique_shirt_results = int(entry_opt_var7.get()) * price_tagW["Pique_shirt"]
        pineJacket_tuple = ("피케 셔츠", str(entry_opt_var7.get()), "30000")
        clothes_quantity_price.append(pineJacket_tuple)
    else:
        Pique_shirt_results = 0

    if check_btn_var8.get() == 1:
        WESTcoat_results = int(entry_opt_var8.get()) * price_tagW["WESTcoat"]
        WESTcoat_tuple = ("정장 조끼", str(entry_opt_var8.get()), "135000")
        clothes_quantity_price.append(WESTcoat_tuple)
    else:
        WESTcoat_results = 0

    if check_btn_var9.get() == 1:
        Tanktop_results = int(entry_opt_var9.get()) * price_tagW["Tanktop"]
        Sleeveless_top_tuple = ("탱크탑", str(entry_opt_var9.get()), "20000")
        clothes_quantity_price.append(Sleeveless_top_tuple)
    else:
        Tanktop_results = 0

    if check_btn_var10.get() == 1:
        Coat_results = int(entry_opt_var10.get()) * price_tagW["Coat"]
        Coat_tuple = ("코트", str(entry_opt_var10.get()), "450000")
        clothes_quantity_price.append(Coat_tuple)
    else:
        Coat_results = 0

    if check_btn_var11.get() == 1:
        Pencidress_results = int(entry_opt_var11.get()) * price_tagW["Pencidress"]
        Pencidress_tuple = ("펜시 드레스", str(entry_opt_var11.get()), "290000")
        clothes_quantity_price.append(Pencidress_results)
    else:
        Pencidress_results = 0

    if check_btn_var12.get() == 1:
        Maxidress_results = int(entry_opt_var12.get()) * price_tagW["Maxidress"]
        Maxidress_tuple = ("멕시드레스", str(entry_opt_var12.get()), "390000")
        clothes_quantity_price.append(Maxidress_tuple)
    else:
        Maxidress_results = 0

    try:
        global total_cost_of_top
        

        total_cost_of_top = (T_shirt_results + Shirt_results + Sleeveless_shirt_results + dress_results + Sweater_results +
                                Cardigan_results + Pique_shirt_results + WESTcoat_results + Tanktop_results +
                                Coat_results + Pencidress_results + Maxidress_results    )

        entry_cost_top.delete(0, END)
        entry_cost_top.insert(0,  str(total_cost_of_top)+" 원. " )
    except NameError:
        error_handling("error")
    except ValueError:
        messagebox.showerror("Error")

    # ================================ bottom-top ==============================

    bottom_price_tagW = {"Jeans": 125000, "Shorts": 105000, "wide_jeans": 100000, "penci_skirt": 150000,
                       "silk_skirt": 290000, "leather_mini_skirt": 115000, "lap_skirt": 79000, "Relaxedfit_jeans": 130000,
                       "skirt_dress": 250000, "mini_dress": 225000, "maxi_dress": 150000, "Belted_dress": 200000}

    if lf2_check_btn_var1.get() == 1:
        wx1 = int(lf2_entry_opt_var1.get()) * bottom_price_tagW["Jeans"]
        Jeans_tuple = ("청바지", str(lf2_entry_opt_var1.get()), "125000")
        clothes_quantity_price.append(Jeans_tuple)
    else:
        wx1 = 0

    if lf2_check_btn_var2.get() == 1:
        x2 = int(lf2_entry_opt_var2.get()) * bottom_price_tagW["Shorts"]
        Shorts_tuple = ("반바지", str(lf2_entry_opt_var2.get()), "105000")
        clothes_quantity_price.append(Shorts_tuple)
    else:
        wx2 = 0

    if lf2_check_btn_var3.get() == 1:
        x3 = int(lf2_entry_opt_var3.get()) * bottom_price_tagW["wide_jeans"]
        wide_jeans_tuple = ("와이드핏 청바지", str(lf2_entry_opt_var3.get()), "100000")
        clothes_quantity_price.append(wide_jeans_tuple)
    else:
        wx3 = 0

    if lf2_check_btn_var4.get() == 1:
        x4 = int(lf2_entry_opt_var4.get()) * bottom_price_tagW["penci_skirt"]
        penci_skirt_tuple = ("펜시 치마", str(lf2_entry_opt_var4.get()), "150000")
        clothes_quantity_price.append(penci_skirt_tuple)
    else:
        wx4 = 0

    if lf2_check_btn_var5.get() == 1:
        x5 = int(lf2_entry_opt_var5.get()) * bottom_price_tagW["silk_skirt"]
        silk_skirt_tuple = ("실크 치마", str(lf2_entry_opt_var5.get()), "290000")
        clothes_quantity_price.append(silk_skirt_tuple)
    else:
        wx5 = 0

    if lf2_check_btn_var6.get() == 1:
        wx6 = int(lf2_entry_opt_var6.get()) * bottom_price_tagW["leather_mini_skirt"]
        leather_mini_skirt_tuple = ("가죽 미니 스커트", str(lf2_entry_opt_var6.get()), "115000")
        clothes_quantity_price.append(leather_mini_skirt_tuple)
    else:
        wx6 = 0

    if lf2_check_btn_var7.get() == 1:
        wx7 = int(lf2_entry_opt_var7.get()) * bottom_price_tagW["lap_skirt"]
        lap_skirt_tuple = ("랩 스커트", str(lf2_entry_opt_var7.get()), "79000")
        clothes_quantity_price.append(lap_skirt_tuple)
    else:
        wx7 = 0

    if lf2_check_btn_var8.get() == 1:
        wx8 = int(lf2_entry_opt_var8.get()) * bottom_price_tagW["Relaxedfit_jeans"]
        Relaxedfit_jeans_tuple = ("릴렉스 핏 청파지", str(lf2_entry_opt_var8.get()), "130000")
        clothes_quantity_price.append(Relaxedfit_jeans_tuple)
    else:
        wx8 = 0

    if lf2_check_btn_var9.get() == 1:
        wx9 = int(lf2_entry_opt_var9.get()) * bottom_price_tagW["skirt_dress"]
        skirt_dress_tuple = ("셔츠 드레스", str(lf2_entry_opt_var9.get()), "250000")
        clothes_quantity_price.append(skirt_dress_tuple)
    else:
        wx9 = 0

    if lf2_check_btn_var10.get() == 1:
        wx10 = int(lf2_entry_opt_var10.get()) * bottom_price_tagW["mini_dress"]
        mini_dress_tuple = ("미니 드레스", str(lf2_entry_opt_var10.get()), "225000")
        clothes_quantity_price.append(mini_dress_tuple)
    else:
        wx10 = 0

    if lf2_check_btn_var11.get() == 1:
        wx11 = int(lf2_entry_opt_var11.get()) * bottom_price_tagW["maxi_dress"]
        maxi_dress_tuple = ("맥시 드레스", str(lf2_entry_opt_var11.get()), "150000")
        clothes_quantity_price.append(maxi_dress_tuple)
    else:
        wx11 = 0

    if lf2_check_btn_var12.get() == 1:
        wx12 = int(lf2_entry_opt_var12.get()) * bottom_price_tagW["Belted_dress"]
        Belted_dress_tuple = ("벨티드 드레스", str(lf2_entry_opt_var12.get()), "200000")
        clothes_quantity_price.append(Belted_dress_tuple)
    else:
        wx12 = 0

    try:
        global wx13
      
        wx13 = 0

        global total_cost_of_bottom
        global tax_on_all_items
        total_cost_of_bottom = (wx1 + wx2 + wx3 + wx4 + wx5 + wx6 + wx7 + wx8 + wx9 + wx10 + wx11 + wx12 + wx13)

        entry_cost_bottom.delete(0, END)
        entry_cost_bottom.insert(0,  str(total_cost_of_bottom)+" 원. ")

        entry_total_cost.delete(0, END)
        entry_total_cost.insert(0,  str(total_cost_of_bottom + total_cost_of_top)+" 원. ")

        tax_on_all_items = 0.10 * (total_cost_of_top + total_cost_of_bottom)

        tax_entry.delete(0, END)
        tax_entry.insert(0,  str(tax_on_all_items)+" 원. ")

        subtotal_entry.delete(0, END)
        subtotal_entry.insert(0,  str(total_cost_of_bottom + total_cost_of_top)+" 원. ")

        total_entry.delete(0, END)
        total_entry.insert(0,  str(total_cost_of_bottom + total_cost_of_top + tax_on_all_items)+" 원. ")
    except NameError:
        error_handling("error!")
    except ValueError:
        messagebox.showerror("Error!", "Price your items first, before finding the total.")

def print_receipt():
    try:
        clothes_field.delete("1.0", END)
        clothes_field.insert(END, "  짱짱 옷가게." + "\n")
        clothes_field.insert(END, " " + "\n")

        x = 0
        while True:
            clothes_field.insert(END, clothes_quantity_price[x][0] + " :   " + clothes_quantity_price[x][1] + " 개당   " +
                                 clothes_quantity_price[x][2]+"원" + "\n")
            x += 1

            if x < len(clothes_quantity_price):
                continue
            else:
                break

        clothes_field.insert(END, " " + "\n")
        clothes_field.insert(END, "상의 가격: " + entry_cost_top_var.get() + "\n")
        clothes_field.insert(END, "하의 가격: " + entry_cost_bottom_var.get() + "\n")
        clothes_field.insert(END, "총 금액:  " + entry_total_cost_var.get() + "\n")
        clothes_field.insert(END, " " + "\n")
        clothes_field.insert(END, " ***** 감사합니다. *****" + "\n")
    except IndexError:
        messagebox.showerror("Error!", "옷을 골라주세요")
    except NameError:
        messagebox.showerror("Error!", "옷을 골라주세요")

def book_keeping():
    # 예약
    try:
        global book
        destination_file = "booking.xlsx"
        book = load_workbook(filename=destination_file)
        sheet = book.active

        row = 2
        column = 1
        global today
        today = time.strftime("%x")
        id_ = random.randint(1, 1000000)

        if clothes_quantity_price:
            while True:
                if sheet.cell(row=row, column=column).value:
                    row += 1
                    continue
                else:
                    sheet.cell(row=row, column=column).value = today
                    column += 1
                    sheet.cell(row=row, column=column).value = id_
                    column += 1

                    # adding elements to the item cell
                    x = len(clothes_quantity_price)
                    sheet.cell(row=row, column=column).value = x
                    column += 1

                    sheet.cell(row=row, column=column).value = total_cost_of_bottom + total_cost_of_top
                    column += 1
                    sheet.cell(row=row, column=column).value = tax_on_all_items

                    if column == 5:
                        break

            book.save(filename=destination_file)
            messagebox.showinfo("예약 완료 !")
        else:
            messagebox.showerror("물건을 넣어주세요")
    except NameError:
        messagebox.showerror("물건을 넣어주세요")

def today_report():
    destination_file2 = "booking.xlsx"
    work_book = load_workbook(filename=destination_file2)

    # GRABBING THE ACTIVE SHEET.
    sheet2 = work_book.active

    row = 2
    column = 1
    max_row = sheet2.max_row
    total_amount = []
    total_vat = []
    today2 = time.strftime("%x")

    x = 0
    while x < max_row:
        if sheet2.cell(row=row, column=column).value == today2:
            total_amount.append(sheet2.cell(row=row, column=4).value)
            total_vat.append(sheet2.cell(row=row, column=5).value)
            row += 1
            x += 1
            continue
        else:
            row += 1
            x += 1
            continue

    answer1 = 0
    for integer in total_amount:
        answer1 += integer

    answer2 = 0
    for float_point in total_vat:
        answer2 += float_point

    clothes_field.delete("1.0", END)
    clothes_field.insert(END, "짱짱 옷가게         " + today2 + "." + "\n")
    clothes_field.insert(END, "  " + "\n") 
    clothes_field.insert(END, "         ***오늘 판매량***" + "\n")
    clothes_field.insert(END, "  " + "\n")  
    clothes_field.insert(END, "전체 판매 금액: " +  str(answer1) + " 원. "+"\n")
    clothes_field.insert(END, "  " + "\n") 
    clothes_field.insert(END, "전체 세금 금액: " +  str(answer2) +" 원. " +"\n")
    clothes_field.insert(END, "  " + "\n")  
    clothes_field.insert(END, "총 수입 + 세금: " +  str(answer1 + answer2) +" 원. " + "\n")
    clothes_field.insert(END, "  " + "\n")  

def highlightFacebyrec(net, frame, conf_threshold=0.7):
    frameOpencvDnn=frame.copy()
    frameHeight=frameOpencvDnn.shape[0]
    frameWidth=frameOpencvDnn.shape[1]
    blob=cv2.dnn.blobFromImage(frameOpencvDnn, 1.0, (300, 300), [104, 117, 123], True, False)

    net.setInput(blob)
    detections=net.forward()
    faceBoxes=[]
    for i in range(detections.shape[2]):
        confidence=detections[0,0,i,2]
        if confidence>conf_threshold:
            x1=int(detections[0,0,i,3]*frameWidth)
            y1=int(detections[0,0,i,4]*frameHeight)
            x2=int(detections[0,0,i,5]*frameWidth)
            y2=int(detections[0,0,i,6]*frameHeight)
            faceBoxes.append([x1,y1,x2,y2])
            cv2.rectangle(frameOpencvDnn, (x1,y1), (x2,y2), (0,0,0), int(round(frameHeight/150)), 8)
    return frameOpencvDnn,faceBoxes

def close_opencv_window():
    cv2.destroyAllWindows()

parser=argparse.ArgumentParser()
parser.add_argument('--image')

args=parser.parse_args()

faceProto="opencv_face_detector.pbtxt"
faceModel="opencv_face_detector_uint8.pb"
ageProto="age_deploy.prototxt"
ageModel="age_net.caffemodel"
genderProto="gender_deploy.prototxt"
genderModel="gender_net.caffemodel"


Sleevless_shirtNum=int(0)

MODEL_MEAN_VALUES=(78.4263377603, 87.7689143744, 114.895847746)


ageList=['(0-3)', '(4-8)', '(8-15)', '(16-20)', '(21-29)', '(30-40)','(41-50)']
genderList=['Male','Female']


faceNet=cv2.dnn.readNet(faceModel,faceProto)
ageNet=cv2.dnn.readNet(ageModel,ageProto)
genderNet=cv2.dnn.readNet(genderModel,genderProto)

video=cv2.VideoCapture(args.image if args.image else 0)
padding=20


#========================================================메인===========================#
while True:
    hasFrame,frame=video.read()    
    resultImg,faceBoxes=highlightFacebyrec(faceNet,frame)   
    for faceBox in faceBoxes:
        start_time = cv2.getTickCount()

        face=frame[max(0,faceBox[1]-padding):
                   min(faceBox[3]+padding,frame.shape[0]-1),max(0,faceBox[0]-padding)
                   :min(faceBox[2]+padding, frame.shape[1]-1)]
        
        blob=cv2.dnn.blobFromImage(face, 1.0, (227,227), MODEL_MEAN_VALUES, swapRB=False)
        genderNet.setInput(blob)
        genderPreds=genderNet.forward()
        gender=genderList[genderPreds[0].argmax()]
        print(f'Gender: {gender}')
        
        ageNet.setInput(blob)
        agePreds=ageNet.forward()
        age=ageList[agePreds[0].argmax()]
        print(f'Age: {age[1:-1]} years')
        
        if gender=='Male':
            close_opencv_window()               
            root = Tk() 
            color = "Light Gray"
            color2 = "Gray"

            #이미지 set
            tsimage = tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\ts.png")
            msimage = tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\ms.png")
            sleevlessimage=tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\sleevlessimage.png")
            suitimage=tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\suitimage.png")
            sweaterimg=tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\sweaterimg.png")
            piqueimg=tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\piqueimg.png")
            cardiganimg=tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\cardiganimg.png")
            vimage = tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\vimage.png")
            capeimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\capeimg.png")
            coatimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\coatimg.png")
            blouseimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\blouseimg.png")
            sweatsuitimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\sweatsuitimg.png")


            #하의 이미지셋
            Jeansimg = tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\Jeansimg.png")
            Shortsimg = tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\Shortsimg.png")
            Straight_wool_trouserimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\Straight_wool_trouserimg.png")
            Dungareeimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\Dungareeimg.png")
            Taylored_shortsimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\Taylored_shortsimg.png")
            Elastic_trouserimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\Elastic_trouserimg.png")
            Towering_shortsimg = tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\Towering_shortsimg.png")
            Relaxedfit_jeansimg = tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\Relaxedfit_jeansimg.png")
            uility_shortsimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\uility_shortsimg.png")
            Seersucker_shortsimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\Seersucker_shortsimg.png")
            cargo_pantsimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\cargo_pantsimg.png")
            swarm_shortsimg= tk.PhotoImage(file="C:\\Users\\cho11\\Desktop\\화이팅\\swarm_shortsimg.png")


            newwindow=tk.Toplevel()
            newwindow.grid
            newwindow.title("사진입니다.")
            newwindow.geometry("300x300")


            image_1label = tk.Label(newwindow, image="") 
            image_1label.pack()   
            image_2label = tk.Label(newwindow, image="") 
            image_2label.pack() 
            image_3label = tk.Label(newwindow, image="") 
            image_3label.pack()   
            image_4label = tk.Label(newwindow, image="") 
            image_4label.pack() 
            image_5label = tk.Label(newwindow, image="") 
            image_5label.pack()   
            image_6label = tk.Label(newwindow, image="") 
            image_6label.pack() 
            image_7label = tk.Label(newwindow, image="") 
            image_7label.pack()   
            image_8label = tk.Label(newwindow, image="") 
            image_8label.pack() 
            image_9label = tk.Label(newwindow, image="") 
            image_9label.pack() 
            image_10label = tk.Label(newwindow, image="") 
            image_10label.pack()   
            image_11label = tk.Label(newwindow, image="") 
            image_11label.pack() 
            image_12label = tk.Label(newwindow, image="") 
            image_12label.pack()   


            bimage_1label = tk.Label(newwindow, image="") 
            bimage_1label.pack()   
            bimage_2label = tk.Label(newwindow, image="") 
            bimage_2label.pack() 
            bimage_3label = tk.Label(newwindow, image="") 
            bimage_3label.pack()   
            bimage_4label = tk.Label(newwindow, image="") 
            bimage_4label.pack() 
            bimage_5label = tk.Label(newwindow, image="") 
            bimage_5label.pack()   
            bimage_6label = tk.Label(newwindow, image="") 
            bimage_6label.pack() 
            bimage_7label = tk.Label(newwindow, image="") 
            bimage_7label.pack()   
            bimage_8label = tk.Label(newwindow, image="") 
            bimage_8label.pack() 
            bimage_9label = tk.Label(newwindow, image="") 
            bimage_9label.pack() 
            bimage_10label = tk.Label(newwindow, image="") 
            bimage_10label.pack()   
            bimage_11label = tk.Label(newwindow, image="") 
            bimage_11label.pack() 
            bimage_12label = tk.Label(newwindow, image="") 
            bimage_12label.pack()   

            root.resizable(width=FALSE, height=FALSE)
            root.geometry("1200x650")
            root.title("For Male SYSTEM")
            
            root.configure(bg=color)
            top_frame = Frame(root, bg=color, width=1200, height=50)
            top_frame.pack(side=TOP)

           
          
            
            bottom_frame = Frame(root, bg=color, width=1200, height=150)
            bottom_frame.pack(side=BOTTOM)
        
            left_frame = Frame(root, bg=color, width=300, height=500)
            left_frame.pack(side=LEFT)

            left_frame2 = Frame(root, bg=color, width=300, height=500)
            left_frame2.pack(side=LEFT, padx=5)

            right_frame = Frame(root, bg=color, width=800, height=150)
            right_frame.place(x=610, y=60)
                     
            right_frame2 = Frame(root, bg=color, width=800, height=280)
            right_frame2.place(x=610, y=210)

            title_label = Label(top_frame, text="짱짱 옷가게", font=("Aisha Latin Semibold", 20, "bold"), bg=color,
                                pady=10)
            title_label.pack(side=LEFT)

            label_cost_top = Label(right_frame, text=" 상의 가격 : ", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            label_cost_top.place(x=10, y=5)

            entry_cost_top_var = StringVar()
            entry_cost_top = Entry(right_frame, textvariable=entry_cost_top_var, width=10)
            entry_cost_top.place(x=160, y=5)


            label_cost_bottom = Label(right_frame, text="하의 가격 : ", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            label_cost_bottom.place(x=10, y=50)

            entry_cost_bottom_var = StringVar()
            entry_cost_bottom = Entry(right_frame, textvariable=entry_cost_bottom_var, width=10)
            entry_cost_bottom.place(x=160, y=50)


            label_total_cost = Label(right_frame, text="총 가격     : ", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            label_total_cost.place(x=10, y=95)

            entry_total_cost_var = StringVar()
            entry_total_cost = Entry(right_frame, textvariable=entry_total_cost_var, width=10)
            entry_total_cost.place(x=160, y=95)


            tax_label = Label(right_frame, text="세금 :", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            tax_label.place(x=350, y=5)

            tax_var = StringVar()
            tax_entry = Entry(right_frame, textvariable=tax_var, width=10)
            tax_entry.place(x=410, y=5)


            subtotal = Label(right_frame, text="가격 :", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            subtotal.place(x=350, y=50)

            subtotal_var = StringVar()
            subtotal_entry = Entry(right_frame, textvariable=subtotal_var, width=10)
            subtotal_entry.place(x=410, y=50)


            total = Label(right_frame, text="총 가격 :", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            total.place(x=350, y=95)

            total_var = StringVar()
            total_entry = Entry(right_frame, textvariable=total_var, width=10)
            total_entry.place(x=410, y=95)

            # ==============================Text field on right_frame2========================
            clothes_field = Text(right_frame2, width=75, height=23)
            clothes_field.pack(side=TOP)

            # ==============================Check-boxes, label, & an entry on the left_frame=====================

            check_btn_var1 = IntVar()
            check_btn1 = Checkbutton(left_frame, text="티셔츠", variable=check_btn_var1,command=show_image)
            check_btn1.place(x=3, y=5)

            check_btn_var2 = IntVar()
            check_btn2 = Checkbutton(left_frame, text="셔츠", variable=check_btn_var2,command=show_image)
            check_btn2.place(x=3, y=35)

            check_btn_var3 = IntVar()
            check_btn3 = Checkbutton(left_frame, text="민소매 ", variable=check_btn_var3,command=show_image)
            check_btn3.place(x=3, y=65)

            check_btn_var4 = IntVar()
            check_btn4 = Checkbutton(left_frame, text="정장", variable=check_btn_var4,command=show_image)
            check_btn4.place(x=3, y=95)

            check_btn_var5 = IntVar()
            check_btn5 = Checkbutton(left_frame, text="스웨터", variable=check_btn_var5,command=show_image)
            check_btn5.place(x=3, y=125)

            check_btn_var6 = IntVar()
            check_btn6 = Checkbutton(left_frame, text="가디건", variable=check_btn_var6,command=show_image)
            check_btn6.place(x=3, y=155)

            check_btn_var7 = IntVar()
            check_btn7 = Checkbutton(left_frame, text="피케 셔츠", variable=check_btn_var7,command=show_image)
            check_btn7.place(x=3, y=185)

            check_btn_var8 = IntVar()
            check_btn8 = Checkbutton(left_frame, text="조끼", variable=check_btn_var8,command=show_image)
            check_btn8.place(x=3, y=215)

            check_btn_var9 = IntVar()
            check_btn9 = Checkbutton(left_frame, text="케이프 코트", variable=check_btn_var9,command=show_image)
            check_btn9.place(x=3, y=245)

            check_btn_var10 = IntVar()
            check_btn10 = Checkbutton(left_frame, text="코트", variable=check_btn_var10,command=show_image)
            check_btn10.place(x=3, y=275)

            check_btn_var11 = IntVar()
            check_btn11 = Checkbutton(left_frame, text="블라우스", variable=check_btn_var11,command=show_image)
            check_btn11.place(x=3, y=305)

            check_btn_var12 = IntVar()
            check_btn12 = Checkbutton(left_frame, text="운동복", variable=check_btn_var12,command=show_image)
            check_btn12.place(x=3, y=335)

          
            # ======================entry-options on the left frame=============================
            list_entry = ["0", "1", "2", "3", "4", "5"]

           
            entry_opt_var1 = StringVar()
            entry_opt_var1.set(list_entry[0])
            entry_opt1 = OptionMenu(left_frame, entry_opt_var1, *list_entry)
            entry_opt1.configure(bg=color, font=("arial", 7))
            entry_opt1.place(x=200, y=5)

            entry_opt_var2 = StringVar()
            entry_opt_var2.set(list_entry[0])
            entry_opt2 = OptionMenu(left_frame, entry_opt_var2, *list_entry)
            entry_opt2.configure(bg=color, font=("arial", 7))
            entry_opt2.place(x=200, y=35)

            entry_opt_var3 = StringVar()
            entry_opt_var3.set(list_entry[0])
            entry_opt3 = OptionMenu(left_frame, entry_opt_var3, *list_entry)
            entry_opt3.configure(bg=color, font=("arial", 7))
            entry_opt3.place(x=200, y=65)

            entry_opt_var4 = StringVar()
            entry_opt_var4.set(list_entry[0])
            entry_opt4 = OptionMenu(left_frame, entry_opt_var4, *list_entry)
            entry_opt4.configure(bg=color, font=("arial", 7))
            entry_opt4.place(x=200, y=95)

            entry_opt_var5 = StringVar()
            entry_opt_var5.set(list_entry[0])
            entry_opt5 = OptionMenu(left_frame, entry_opt_var5, *list_entry)
            entry_opt5.configure(bg=color, font=("arial", 7))
            entry_opt5.place(x=200, y=125)

            entry_opt_var6 = StringVar()
            entry_opt_var6.set(list_entry[0])
            entry_opt6 = OptionMenu(left_frame, entry_opt_var6, *list_entry)
            entry_opt6.configure(bg=color, font=("arial", 7))
            entry_opt6.place(x=200, y=155)

            entry_opt_var7 = StringVar()
            entry_opt_var7.set(list_entry[0])
            entry_opt7 = OptionMenu(left_frame, entry_opt_var7, *list_entry)
            entry_opt7.configure(bg=color, font=("arial", 7))
            entry_opt7.place(x=200, y=185)

            entry_opt_var8 = StringVar()
            entry_opt_var8.set(list_entry[0])
            entry_opt8 = OptionMenu(left_frame, entry_opt_var8, *list_entry)
            entry_opt8.configure(bg=color, font=("arial", 7))
            entry_opt8.place(x=200, y=215)

            entry_opt_var9 = StringVar()
            entry_opt_var9.set(list_entry[0])
            entry_opt9 = OptionMenu(left_frame, entry_opt_var9, *list_entry)
            entry_opt9.configure(bg=color, font=("arial", 7))
            entry_opt9.place(x=200, y=245)

            entry_opt_var10 = StringVar()
            entry_opt_var10.set(list_entry[0])
            entry_opt10 = OptionMenu(left_frame, entry_opt_var10, *list_entry)
            entry_opt10.configure(bg=color, font=("arial", 7))
            entry_opt10.place(x=200, y=275)

            entry_opt_var11 = StringVar()
            entry_opt_var11.set(list_entry[0])
            entry_opt11 = OptionMenu(left_frame, entry_opt_var11, *list_entry)
            entry_opt11.configure(bg=color, font=("arial", 7))
            entry_opt11.place(x=200, y=305)

            entry_opt_var12 = StringVar()
            entry_opt_var12.set(list_entry[0])
            entry_opt12 = OptionMenu(left_frame, entry_opt_var12, *list_entry)
            entry_opt12.configure(bg=color, font=("arial", 7))
            entry_opt12.place(x=200, y=335)


            lf2_check_btn_var1 = IntVar()
            lf2_check_btn1 = Checkbutton(left_frame2, text="청바지", variable=lf2_check_btn_var1,command=show_image)
            lf2_check_btn1.place(x=3, y=5)

            lf2_check_btn_var2 = IntVar()
            lf2_check_btn2 = Checkbutton(left_frame2, text="반바지", variable=lf2_check_btn_var2,command=show_image)
            lf2_check_btn2.place(x=3, y=35)

            lf2_check_btn_var3 = IntVar()
            lf2_check_btn3 = Checkbutton(left_frame2, text="울바지", variable=lf2_check_btn_var3,command=show_image)
            lf2_check_btn3.place(x=3, y=65)

            lf2_check_btn_var4 = IntVar()
            lf2_check_btn4 = Checkbutton(left_frame2, text="멜빵바지", variable=lf2_check_btn_var4,command=show_image)
            lf2_check_btn4.place(x=3, y=95)

            lf2_check_btn_var5 = IntVar()
            lf2_check_btn5 = Checkbutton(left_frame2, text="정장 바지", variable=lf2_check_btn_var5,command=show_image)
            lf2_check_btn5.place(x=3, y=125)

            lf2_check_btn_var6 = IntVar()
            lf2_check_btn6 = Checkbutton(left_frame2, text="기능성 바지", variable=lf2_check_btn_var6,command=show_image)
            lf2_check_btn6.place(x=3, y=155)

            lf2_check_btn_var7 = IntVar()
            lf2_check_btn7 = Checkbutton(left_frame2, text="타월 바지", variable=lf2_check_btn_var7,command=show_image)
            lf2_check_btn7.place(x=3, y=185)

            lf2_check_btn_var8 = IntVar()
            lf2_check_btn8 = Checkbutton(left_frame2, text="릭렉스 핏 청바지", variable=lf2_check_btn_var8,command=show_image)
            lf2_check_btn8.place(x=3, y=215)

            lf2_check_btn_var9 = IntVar()
            lf2_check_btn9 = Checkbutton(left_frame2, text="운동 바지", variable=lf2_check_btn_var9,command=show_image)
            lf2_check_btn9.place(x=3, y=245)

            lf2_check_btn_var10 = IntVar()
            lf2_check_btn10 = Checkbutton(left_frame2, text="시어서커 반바지", variable=lf2_check_btn_var10,command=show_image)
            lf2_check_btn10.place(x=3, y=275)

            lf2_check_btn_var11 = IntVar()
            lf2_check_btn11 = Checkbutton(left_frame2, text="카고바지", variable=lf2_check_btn_var11,command=show_image)
            lf2_check_btn11.place(x=3, y=305)

            lf2_check_btn_var12 = IntVar()
            lf2_check_btn12 = Checkbutton(left_frame2, text="수영복 바지", variable=lf2_check_btn_var12,command=show_image)
            lf2_check_btn12.place(x=3, y=335)

            lf2_list_entry = ["0", "1", "2", "3", "4", "5"]

            lf2_entry_opt_var1 = StringVar()
            lf2_entry_opt_var1.set(list_entry[0])
            lf2_entry_opt1 = OptionMenu(left_frame2, lf2_entry_opt_var1, *lf2_list_entry)
            lf2_entry_opt1.configure(bg=color, font=("arial", 7))
            lf2_entry_opt1.place(x=200, y=5)

            lf2_entry_opt_var2 = StringVar()
            lf2_entry_opt_var2.set(list_entry[0])
            lf2_entry_opt2 = OptionMenu(left_frame2, lf2_entry_opt_var2, *lf2_list_entry)
            lf2_entry_opt2.configure(bg=color, font=("arial", 7))
            lf2_entry_opt2.place(x=200, y=35)

            lf2_entry_opt_var3 = StringVar()
            lf2_entry_opt_var3.set(list_entry[0])
            lf2_entry_opt3 = OptionMenu(left_frame2, lf2_entry_opt_var3, *lf2_list_entry)
            lf2_entry_opt3.configure(bg=color, font=("arial", 7))
            lf2_entry_opt3.place(x=200, y=65)

            lf2_entry_opt_var4 = StringVar()
            lf2_entry_opt_var4.set(list_entry[0])
            lf2_entry_opt4 = OptionMenu(left_frame2, lf2_entry_opt_var4, *lf2_list_entry)
            lf2_entry_opt4.configure(bg=color, font=("arial", 7))
            lf2_entry_opt4.place(x=200, y=95)

            lf2_entry_opt_var5 = StringVar()
            lf2_entry_opt_var5.set(list_entry[0])
            lf2_entry_opt5 = OptionMenu(left_frame2, lf2_entry_opt_var5, *lf2_list_entry)
            lf2_entry_opt5.configure(bg=color, font=("arial", 7))
            lf2_entry_opt5.place(x=200, y=125)

            lf2_entry_opt_var6 = StringVar()
            lf2_entry_opt_var6.set(list_entry[0])
            lf2_entry_opt6 = OptionMenu(left_frame2, lf2_entry_opt_var6, *lf2_list_entry)
            lf2_entry_opt6.configure(bg=color, font=("arial", 7))
            lf2_entry_opt6.place(x=200, y=155)

            lf2_entry_opt_var7 = StringVar()
            lf2_entry_opt_var7.set(list_entry[0])
            lf2_entry_opt7 = OptionMenu(left_frame2, lf2_entry_opt_var7, *lf2_list_entry)
            lf2_entry_opt7.configure(bg=color, font=("arial", 7))
            lf2_entry_opt7.place(x=200, y=185)

            lf2_entry_opt_var8 = StringVar()
            lf2_entry_opt_var8.set(list_entry[0])
            lf2_entry_opt8 = OptionMenu(left_frame2, lf2_entry_opt_var8, *lf2_list_entry)
            lf2_entry_opt8.configure(bg=color, font=("arial", 7))
            lf2_entry_opt8.place(x=200, y=215)

            lf2_entry_opt_var9 = StringVar()
            lf2_entry_opt_var9.set(list_entry[0])
            lf2_entry_opt9 = OptionMenu(left_frame2, lf2_entry_opt_var9, *lf2_list_entry)
            lf2_entry_opt9.configure(bg=color, font=("arial", 7))
            lf2_entry_opt9.place(x=200, y=245)

            lf2_entry_opt_var10 = StringVar()
            lf2_entry_opt_var10.set(list_entry[0])
            lf2_entry_opt10 = OptionMenu(left_frame2, lf2_entry_opt_var10, *lf2_list_entry)
            lf2_entry_opt10.configure(bg=color, font=("arial", 7))
            lf2_entry_opt10.place(x=200, y=275)

            lf2_entry_opt_var11 = StringVar()
            lf2_entry_opt_var11.set(list_entry[0])
            lf2_entry_opt11 = OptionMenu(left_frame2, lf2_entry_opt_var11, *lf2_list_entry)
            lf2_entry_opt11.configure(bg=color, font=("arial", 7))
            lf2_entry_opt11.place(x=200, y=305)

            lf2_entry_opt_var12 = StringVar()
            lf2_entry_opt_var12.set(list_entry[0])
            lf2_entry_opt12 = OptionMenu(left_frame2, lf2_entry_opt_var12, *lf2_list_entry)
            lf2_entry_opt12.configure(bg=color, font=("arial", 7))
            lf2_entry_opt12.place(x=200, y=335)

          

            total_btn_btm_frame = Button(bottom_frame, text="주문", width=60, bg=color2, command=lambda: total_salesM())
            total_btn_btm_frame.place(x=5, y=60)

            receipt_btn_btm_frame = Button(bottom_frame, text="영수증", width=60, bg=color2, command=lambda: print_receipt())
            receipt_btn_btm_frame.place(x=400, y=60)

            reset_btn_btm_frame = Button(bottom_frame, text="처음 화면", width=60, bg=color2, command=lambda: reset_sales())
            reset_btn_btm_frame.place(x=795, y=60)

            book_k_btn_btm_frame = Button(bottom_frame, text="예약", width=60, bg=color2, command=lambda: book_keeping())
            book_k_btn_btm_frame.place(x=5, y=115)

            report_btn_btm_frame = Button(bottom_frame, text="오늘 판매량", width=60, bg=color2, command=lambda: today_report())
            report_btn_btm_frame.place(x=400, y=115)

            exit_btn_btm_frame = Button(bottom_frame, text="종료", width=60, bg="red", command=lambda: exit_program())
            exit_btn_btm_frame.place(x=795, y=115)        
            root.mainloop()


            ##########################female
        elif gender == 'Female':
            root = Tk()
            root.resizable(width=FALSE, height=FALSE)
            root.geometry("1200x650")
            root.title("For Female SYSTEM")
            color = "Light Gray"
            color2 = "Gray"
            root.configure(bg=color)
            top_frame = Frame(root, bg=color, width=1200, height=50)
            top_frame.pack(side=TOP)
           

            
            bottom_frame = Frame(root, bg=color, width=1200, height=150)
            bottom_frame.pack(side=BOTTOM)
        
            left_frame = Frame(root, bg=color, width=300, height=500)
            left_frame.pack(side=LEFT)

            left_frame2 = Frame(root, bg=color, width=300, height=500)
            left_frame2.pack(side=LEFT, padx=5)

            right_frame = Frame(root, bg=color, width=800, height=150)
            right_frame.place(x=610, y=60)
                     
            right_frame2 = Frame(root, bg=color, width=800, height=280)
            right_frame2.place(x=610, y=210)

            title_label = Label(top_frame, text="짱짱 옷가게", font=("Aisha Latin Semibold", 20, "bold"), bg=color,
                                pady=10)
            title_label.pack(side=LEFT)

            label_cost_top = Label(right_frame, text=" 상의 가격 : ", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            label_cost_top.place(x=10, y=5)

            entry_cost_top_var = StringVar()
            entry_cost_top = Entry(right_frame, textvariable=entry_cost_top_var, width=10)
            entry_cost_top.place(x=160, y=5)


            label_cost_bottom = Label(right_frame, text="하의 가격 : ", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            label_cost_bottom.place(x=10, y=50)

            entry_cost_bottom_var = StringVar()
            entry_cost_bottom = Entry(right_frame, textvariable=entry_cost_bottom_var, width=10)
            entry_cost_bottom.place(x=160, y=50)


            label_total_cost = Label(right_frame, text="총 가격     : ", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            label_total_cost.place(x=10, y=95)

            entry_total_cost_var = StringVar()
            entry_total_cost = Entry(right_frame, textvariable=entry_total_cost_var, width=10)
            entry_total_cost.place(x=160, y=95)


            tax_label = Label(right_frame, text="세금 :", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            tax_label.place(x=350, y=5)

            tax_var = StringVar()
            tax_entry = Entry(right_frame, textvariable=tax_var, width=10)
            tax_entry.place(x=410, y=5)


            subtotal = Label(right_frame, text="가격 :", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            subtotal.place(x=350, y=50)

            subtotal_var = StringVar()
            subtotal_entry = Entry(right_frame, textvariable=subtotal_var, width=10)
            subtotal_entry.place(x=410, y=50)


            total = Label(right_frame, text="총 가격 :", font=("Aisha Latin Semibold", 10, "bold"), bg=color)
            total.place(x=350, y=95)

            total_var = StringVar()
            total_entry = Entry(right_frame, textvariable=total_var, width=10)
            total_entry.place(x=410, y=95)

            # ==============================Text field on right_frame2========================
            clothes_field = Text(right_frame2, width=58, height=14)
            clothes_field.pack(side=TOP)

            # ==============================Check-boxes, label, & an entry on the left_frame=====================

            check_btn_var1 = IntVar()
            check_btn1 = Checkbutton(left_frame, text="티셔츠", variable=check_btn_var1)
            check_btn1.place(x=3, y=5)

            check_btn_var2 = IntVar()
            check_btn2 = Checkbutton(left_frame, text="셔츠", variable=check_btn_var2)
            check_btn2.place(x=3, y=35)

            check_btn_var3 = IntVar()
            check_btn3 = Checkbutton(left_frame, text="민소매 셔츠", variable=check_btn_var3)
            check_btn3.place(x=3, y=65)

            check_btn_var4 = IntVar()
            check_btn4 = Checkbutton(left_frame, text="드레스", variable=check_btn_var4)
            check_btn4.place(x=3, y=95)

            check_btn_var5 = IntVar()
            check_btn5 = Checkbutton(left_frame, text="스웨터", variable=check_btn_var5)
            check_btn5.place(x=3, y=125)

            check_btn_var6 = IntVar()
            check_btn6 = Checkbutton(left_frame, text="가디건", variable=check_btn_var6)
            check_btn6.place(x=3, y=155)

            check_btn_var7 = IntVar()
            check_btn7 = Checkbutton(left_frame, text="피케 셔츠", variable=check_btn_var7)
            check_btn7.place(x=3, y=185)

            check_btn_var8 = IntVar()
            check_btn8 = Checkbutton(left_frame, text="정장 조끼", variable=check_btn_var8)
            check_btn8.place(x=3, y=215)

            check_btn_var9 = IntVar()
            check_btn9 = Checkbutton(left_frame, text="탱크탑", variable=check_btn_var9)
            check_btn9.place(x=3, y=245)

            check_btn_var10 = IntVar()
            check_btn10 = Checkbutton(left_frame, text="코트", variable=check_btn_var10)
            check_btn10.place(x=3, y=275)

            check_btn_var11 = IntVar()
            check_btn11 = Checkbutton(left_frame, text="펜시 드레스", variable=check_btn_var11)
            check_btn11.place(x=3, y=305)

            check_btn_var12 = IntVar()
            check_btn12 = Checkbutton(left_frame, text="맥시 드레스", variable=check_btn_var12)
            check_btn12.place(x=3, y=335)
            list_entry = ["0", "1", "2", "3", "4", "5"]

           
            entry_opt_var1 = StringVar()
            entry_opt_var1.set(list_entry[0])
            entry_opt1 = OptionMenu(left_frame, entry_opt_var1, *list_entry)
            entry_opt1.configure(bg=color, font=("arial", 7))
            entry_opt1.place(x=200, y=5)

            entry_opt_var2 = StringVar()
            entry_opt_var2.set(list_entry[0])
            entry_opt2 = OptionMenu(left_frame, entry_opt_var2, *list_entry)
            entry_opt2.configure(bg=color, font=("arial", 7))
            entry_opt2.place(x=200, y=35)

            entry_opt_var3 = StringVar()
            entry_opt_var3.set(list_entry[0])
            entry_opt3 = OptionMenu(left_frame, entry_opt_var3, *list_entry)
            entry_opt3.configure(bg=color, font=("arial", 7))
            entry_opt3.place(x=200, y=65)

            entry_opt_var4 = StringVar()
            entry_opt_var4.set(list_entry[0])
            entry_opt4 = OptionMenu(left_frame, entry_opt_var4, *list_entry)
            entry_opt4.configure(bg=color, font=("arial", 7))
            entry_opt4.place(x=200, y=95)

            entry_opt_var5 = StringVar()
            entry_opt_var5.set(list_entry[0])
            entry_opt5 = OptionMenu(left_frame, entry_opt_var5, *list_entry)
            entry_opt5.configure(bg=color, font=("arial", 7))
            entry_opt5.place(x=200, y=125)

            entry_opt_var6 = StringVar()
            entry_opt_var6.set(list_entry[0])
            entry_opt6 = OptionMenu(left_frame, entry_opt_var6, *list_entry)
            entry_opt6.configure(bg=color, font=("arial", 7))
            entry_opt6.place(x=200, y=155)

            entry_opt_var7 = StringVar()
            entry_opt_var7.set(list_entry[0])
            entry_opt7 = OptionMenu(left_frame, entry_opt_var7, *list_entry)
            entry_opt7.configure(bg=color, font=("arial", 7))
            entry_opt7.place(x=200, y=185)

            entry_opt_var8 = StringVar()
            entry_opt_var8.set(list_entry[0])
            entry_opt8 = OptionMenu(left_frame, entry_opt_var8, *list_entry)
            entry_opt8.configure(bg=color, font=("arial", 7))
            entry_opt8.place(x=200, y=215)

            entry_opt_var9 = StringVar()
            entry_opt_var9.set(list_entry[0])
            entry_opt9 = OptionMenu(left_frame, entry_opt_var9, *list_entry)
            entry_opt9.configure(bg=color, font=("arial", 7))
            entry_opt9.place(x=200, y=245)

            entry_opt_var10 = StringVar()
            entry_opt_var10.set(list_entry[0])
            entry_opt10 = OptionMenu(left_frame, entry_opt_var10, *list_entry)
            entry_opt10.configure(bg=color, font=("arial", 7))
            entry_opt10.place(x=200, y=275)

            entry_opt_var11 = StringVar()
            entry_opt_var11.set(list_entry[0])
            entry_opt11 = OptionMenu(left_frame, entry_opt_var11, *list_entry)
            entry_opt11.configure(bg=color, font=("arial", 7))
            entry_opt11.place(x=200, y=305)

            entry_opt_var12 = StringVar()
            entry_opt_var12.set(list_entry[0])
            entry_opt12 = OptionMenu(left_frame, entry_opt_var12, *list_entry)
            entry_opt12.configure(bg=color, font=("arial", 7))
            entry_opt12.place(x=200, y=335)

    

            lf2_check_btn_var1 = IntVar()
            lf2_check_btn1 = Checkbutton(left_frame2, text="청바지", variable=lf2_check_btn_var1)
            lf2_check_btn1.place(x=3, y=5)

            lf2_check_btn_var2 = IntVar()
            lf2_check_btn2 = Checkbutton(left_frame2, text="반바지", variable=lf2_check_btn_var2)
            lf2_check_btn2.place(x=3, y=35)

            lf2_check_btn_var3 = IntVar()
            lf2_check_btn3 = Checkbutton(left_frame2, text="와이드핏 청바지", variable=lf2_check_btn_var3)
            lf2_check_btn3.place(x=3, y=65)

            lf2_check_btn_var4 = IntVar()
            lf2_check_btn4 = Checkbutton(left_frame2, text="펜시 치마", variable=lf2_check_btn_var4)
            lf2_check_btn4.place(x=3, y=95)

            lf2_check_btn_var5 = IntVar()
            lf2_check_btn5 = Checkbutton(left_frame2, text="실크 치마", variable=lf2_check_btn_var5)
            lf2_check_btn5.place(x=3, y=125)

            lf2_check_btn_var6 = IntVar()
            lf2_check_btn6 = Checkbutton(left_frame2, text="가죽 미니 스커트", variable=lf2_check_btn_var6)
            lf2_check_btn6.place(x=3, y=155)

            lf2_check_btn_var7 = IntVar()
            lf2_check_btn7 = Checkbutton(left_frame2, text="랩 스커트", variable=lf2_check_btn_var7)
            lf2_check_btn7.place(x=3, y=185)

            lf2_check_btn_var8 = IntVar()
            lf2_check_btn8 = Checkbutton(left_frame2, text="릴랙스드 핏 청바지", variable=lf2_check_btn_var8)
            lf2_check_btn8.place(x=3, y=215)

            lf2_check_btn_var9 = IntVar()
            lf2_check_btn9 = Checkbutton(left_frame2, text="셔츠 드레스", variable=lf2_check_btn_var9)
            lf2_check_btn9.place(x=3, y=245)

            lf2_check_btn_var10 = IntVar()
            lf2_check_btn10 = Checkbutton(left_frame2, text="미니 드레스", variable=lf2_check_btn_var10)
            lf2_check_btn10.place(x=3, y=275)

            lf2_check_btn_var11 = IntVar()
            lf2_check_btn11 = Checkbutton(left_frame2, text="맥시 드레스", variable=lf2_check_btn_var11)
            lf2_check_btn11.place(x=3, y=305)

            lf2_check_btn_var12 = IntVar()
            lf2_check_btn12 = Checkbutton(left_frame2, text="벨티드 드레스", variable=lf2_check_btn_var12)
            lf2_check_btn12.place(x=3, y=335)

            lf2_list_entry = ["0", "1", "2", "3", "4", "5"]

            lf2_entry_opt_var1 = StringVar()
            lf2_entry_opt_var1.set(list_entry[0])
            lf2_entry_opt1 = OptionMenu(left_frame2, lf2_entry_opt_var1, *lf2_list_entry)
            lf2_entry_opt1.configure(bg=color, font=("arial", 7))
            lf2_entry_opt1.place(x=200, y=5)

            lf2_entry_opt_var2 = StringVar()
            lf2_entry_opt_var2.set(list_entry[0])
            lf2_entry_opt2 = OptionMenu(left_frame2, lf2_entry_opt_var2, *lf2_list_entry)
            lf2_entry_opt2.configure(bg=color, font=("arial", 7))
            lf2_entry_opt2.place(x=200, y=35)

            lf2_entry_opt_var3 = StringVar()
            lf2_entry_opt_var3.set(list_entry[0])
            lf2_entry_opt3 = OptionMenu(left_frame2, lf2_entry_opt_var3, *lf2_list_entry)
            lf2_entry_opt3.configure(bg=color, font=("arial", 7))
            lf2_entry_opt3.place(x=200, y=65)

            lf2_entry_opt_var4 = StringVar()
            lf2_entry_opt_var4.set(list_entry[0])
            lf2_entry_opt4 = OptionMenu(left_frame2, lf2_entry_opt_var4, *lf2_list_entry)
            lf2_entry_opt4.configure(bg=color, font=("arial", 7))
            lf2_entry_opt4.place(x=200, y=95)

            lf2_entry_opt_var5 = StringVar()
            lf2_entry_opt_var5.set(list_entry[0])
            lf2_entry_opt5 = OptionMenu(left_frame2, lf2_entry_opt_var5, *lf2_list_entry)
            lf2_entry_opt5.configure(bg=color, font=("arial", 7))
            lf2_entry_opt5.place(x=200, y=125)

            lf2_entry_opt_var6 = StringVar()
            lf2_entry_opt_var6.set(list_entry[0])
            lf2_entry_opt6 = OptionMenu(left_frame2, lf2_entry_opt_var6, *lf2_list_entry)
            lf2_entry_opt6.configure(bg=color, font=("arial", 7))
            lf2_entry_opt6.place(x=200, y=155)

            lf2_entry_opt_var7 = StringVar()
            lf2_entry_opt_var7.set(list_entry[0])
            lf2_entry_opt7 = OptionMenu(left_frame2, lf2_entry_opt_var7, *lf2_list_entry)
            lf2_entry_opt7.configure(bg=color, font=("arial", 7))
            lf2_entry_opt7.place(x=200, y=185)

            lf2_entry_opt_var8 = StringVar()
            lf2_entry_opt_var8.set(list_entry[0])
            lf2_entry_opt8 = OptionMenu(left_frame2, lf2_entry_opt_var8, *lf2_list_entry)
            lf2_entry_opt8.configure(bg=color, font=("arial", 7))
            lf2_entry_opt8.place(x=200, y=215)

            lf2_entry_opt_var9 = StringVar()
            lf2_entry_opt_var9.set(list_entry[0])
            lf2_entry_opt9 = OptionMenu(left_frame2, lf2_entry_opt_var9, *lf2_list_entry)
            lf2_entry_opt9.configure(bg=color, font=("arial", 7))
            lf2_entry_opt9.place(x=200, y=245)

            lf2_entry_opt_var10 = StringVar()
            lf2_entry_opt_var10.set(list_entry[0])
            lf2_entry_opt10 = OptionMenu(left_frame2, lf2_entry_opt_var10, *lf2_list_entry)
            lf2_entry_opt10.configure(bg=color, font=("arial", 7))
            lf2_entry_opt10.place(x=200, y=275)

            lf2_entry_opt_var11 = StringVar()
            lf2_entry_opt_var11.set(list_entry[0])
            lf2_entry_opt11 = OptionMenu(left_frame2, lf2_entry_opt_var11, *lf2_list_entry)
            lf2_entry_opt11.configure(bg=color, font=("arial", 7))
            lf2_entry_opt11.place(x=200, y=305)

            lf2_entry_opt_var12 = StringVar()
            lf2_entry_opt_var12.set(list_entry[0])
            lf2_entry_opt12 = OptionMenu(left_frame2, lf2_entry_opt_var12, *lf2_list_entry)
            lf2_entry_opt12.configure(bg=color, font=("arial", 7))
            lf2_entry_opt12.place(x=200, y=335)

          
   

            total_btn_btm_frame = Button(bottom_frame, text="주문", width=60, bg=color2, command=lambda: total_salesM())
            total_btn_btm_frame.place(x=5, y=60)

            receipt_btn_btm_frame = Button(bottom_frame, text="영수증", width=60, bg=color2, command=lambda: print_receipt())
            receipt_btn_btm_frame.place(x=400, y=60)

            reset_btn_btm_frame = Button(bottom_frame, text="처음 화면", width=60, bg=color2, command=lambda: reset_sales())
            reset_btn_btm_frame.place(x=795, y=60)

            book_k_btn_btm_frame = Button(bottom_frame, text="예약", width=60, bg=color2, command=lambda: book_keeping())
            book_k_btn_btm_frame.place(x=5, y=115)

            report_btn_btm_frame = Button(bottom_frame, text="오늘 판매량", width=60, bg=color2, command=lambda: today_report())
            report_btn_btm_frame.place(x=400, y=115)

            exit_btn_btm_frame = Button(bottom_frame, text="종료", width=60, bg=color2, command=lambda: exit_program())
            exit_btn_btm_frame.place(x=795, y=115)        
            root.mainloop()
